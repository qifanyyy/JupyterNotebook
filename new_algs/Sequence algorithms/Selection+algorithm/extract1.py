from bs4 import BeautifulSoup
from openpyxl import load_workbook
import os,time
import subprocess
from daemon import DaemonContext
import urllib2,sys
urls = ["http://cloud-computing.softwareinsider.com/l/265/Atlantic-Net",
"http://cloud-computing.softwareinsider.com/l/162/CloudSigma-Holding-AG",
"http://cloud-computing.softwareinsider.com/l/18/Engine-Yard-Inc",
"http://cloud-computing.softwareinsider.com/l/17/Microsoft-Corporation",
"http://cloud-computing.softwareinsider.com/l/150/Lunacloud-Lda",
"http://cloud-computing.softwareinsider.com/l/96/Dimension-Data",
"http://cloud-computing.softwareinsider.com/l/14/Rackspace-US-Inc",
"http://cloud-computing.softwareinsider.com/l/181/iWeb-Technologies-Inc",
"http://cloud-computing.softwareinsider.com/l/262/FireHost-Inc",
"http://cloud-computing.softwareinsider.com/l/307/Devart-S-R-O",
"http://cloud-computing.softwareinsider.com/l/292/DigitalOcean-Inc",
"http://cloud-computing.softwareinsider.com/l/300/Google-Inc",
"http://cloud-computing.softwareinsider.com/l/2/GoGrid-LLC",
"http://cloud-computing.softwareinsider.com/l/261/Selectel-Ltd",
"http://cloud-computing.softwareinsider.com/l/6/Joyent-Inc",
"http://cloud-computing.softwareinsider.com/l/266/ArubaCloud",
"http://cloud-computing.softwareinsider.com/l/270/Nimbix-Inc",
"http://cloud-computing.softwareinsider.com/l/15/Google-App-Engine",
"http://cloud-computing.softwareinsider.com/l/271/Serversaurus",
"http://cloud-computing.softwareinsider.com/l/276/Amazon-Web-Services-Inc",
"http://cloud-computing.softwareinsider.com/l/277/Clever-Cloud-SAS",
"http://cloud-computing.softwareinsider.com/l/25/AT-And-T-Inc",
"http://cloud-computing.softwareinsider.com/l/281/Heroku-Inc",
"http://cloud-computing.softwareinsider.com/l/282/Hewlett-Packard-Development-Company-L-P",
"http://cloud-computing.softwareinsider.com/l/283/Modulus",
"http://cloud-computing.softwareinsider.com/l/285/SoftLayer-Technologies-Inc",
"http://cloud-computing.softwareinsider.com/l/286/Red-Hat-Inc",
"http://cloud-computing.softwareinsider.com/l/291/Ninefold-Pty-Limited",
"http://cloud-computing.softwareinsider.com/l/298/Cloudways-Ltd",
"http://cloud-computing.softwareinsider.com/l/53/ElasticHosts-Ltd",
"http://cloud-computing.softwareinsider.com/l/60/Hewlett-Packard-Development-Company",
"http://cloud-computing.softwareinsider.com/l/317/Zetta-IO-Technology-AS",
"http://cloud-computing.softwareinsider.com/l/321/Stack-Harbor",
"http://cloud-computing.softwareinsider.com/l/334/ESDS-Software-Solution-Pvt-Ltd",
"http://cloud-computing.softwareinsider.com/l/337/SecureRack",
"http://cloud-computing.softwareinsider.com/l/356/Digital-Edge",
"http://cloud-computing.softwareinsider.com/l/358/Linode",
"http://cloud-computing.softwareinsider.com/l/149/Logicworks",
"http://cloud-computing.softwareinsider.com/l/164/Ultimum-Technologies-s-r-o",
"http://cloud-computing.softwareinsider.com/l/169/CloudCentral",
"http://cloud-computing.softwareinsider.com/l/188/Fujitsu",
"http://cloud-computing.softwareinsider.com/l/208/Netmagic-Solutions-Pvt-Ltd",
]
def extract():
    wb = load_workbook(filename = 'TabulatedSLA.xlsx') 
    ws = wb['Sheet1']
    for url in urls:
        try:
            req = urllib2.Request(url, headers={'User-Agent' : 'Magic Browser'})
            con = urllib2.urlopen(req)
            page = con.read()
            soup = BeautifulSoup(page,"html.parser")
            section = soup.find('section',{'data-section-id':'2'})
            rating = soup.find('div',{'class':'ur-avg'})
            headings = section.findAll('div',{'class':'dd-overview-label'})
            values = section.findAll('noscript')
            #print headings,values
            
            sname = url.split('/')[-1]

            print "-----"
            print sname,"\n"
            
            for r in range(1,len(ws.rows)):
                if(ws.cell(row=r,column=3).value == sname):
                    if rating == None:
                        ws.cell(row=r,column=14).value = 0
                    else:
                        ws.cell(row=r,column=14).value = float(rating.text)
                    for col in range(1,len(ws.rows[r])):
                        for i in range(len(headings)):
                            if(headings[i].text == 'Base Plan Price' and ws.cell(row=1,column=col).value == headings[i].text ):
                                print "Updating Base Plan Price"
                                temp = values[i].text.replace("per hour","")
                                ws.cell(row=r,column=col).value = float(temp.replace("$",""))

                            if(headings[i].text == 'Virtual CPU Cores' and ws.cell(row=1,column=col).value == headings[i].text):
                                print "Updating Virtual Cores"
                                ws.cell(row=r,column=col).value = int(values[i].text.replace("vCPU's",""))

                            if(headings[i].text == 'RAM' and ws.cell(row=1,column=col).value == headings[i].text):
                                print "Updating RAM"
                                ws.cell(row=r,column=col).value = int(values[i].text.replace("MB",""))

                            if(headings[i].text == 'Disk Space' and ws.cell(row=1,column=col).value == headings[i].text):
                                print "Updating Disk Space"
                                ws.cell(row=r,column=col).value = int(values[i].text.replace("GB",""))

            print "-----"
            #time.sleep(2)
        except:
            print url

    print "Update Finished"

    wb.save('TabulatedSLA.xlsx')

with daemon.DaemonContext():
    extract()
# import win32api, win32gui, win32con

# os.system("start excel TabulatedSLA.xlsx")
# os.system("taskkill /IM excel.exe")

# time.sleep(5)

# window = win32gui.GetForegroundWindow()
# hwnd = win32gui.FindWindowEx(window, 0, "static", None)
# hbutton = win32gui.FindWindowEx(hwnd, 0, "Button", "Save")

# # mouse button click on the OK button, WM_COMMAND may work too
# win32api.PostMessage(hbutton, win32con.WM_LBUTTONDOWN, 0, 0)
# win32api.PostMessage(hbutton, win32con.WM_LBUTTONUP, 0, 0)


#os.system("initialize.py 1")
#subprocess.Popen("initialize.py 1", shell=True)