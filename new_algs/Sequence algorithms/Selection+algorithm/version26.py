# -*- coding: utf-8 -*-
"""
Created on Sat Mar 16 15:14:21 2019
@author: wmy
"""

import cv2
import numpy as np
import matplotlib.pyplot as plt
from sklearn.cluster import KMeans
import tkinter as tk
from tkinter.filedialog import askopenfilename
from tkinter.filedialog import asksaveasfilename
from tkinter import messagebox
from PIL import Image, ImageTk
from sklearn.cluster import DBSCAN
import operator
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, Font, Alignment

np.random.seed(1)

# std colors
std_colors = {'DG0546':[27.818, -8.375, 8.282],
              'DG0543':[27.818, -5.331, 10.061],
              'DG0647':[30.4, -5.075, 5.519],
              'DG0642':[30.4, -7.885, 8.605],
              'DG0648':[30.4, -5.77, 11.405],
              'DG0746':[32.963, -7.188, 8.464],
              'DG0744':[32.963, -6.138, 12.688],
              'DG0750':[32.963, -8.569, 15.102],
              'MG0851':[35.538, -11.055, 12.566],
              'MG0841':[35.538, -7.952, 14.866],
              'MG0850':[35.538, -9.978, 17.898],
              'MG1050':[38.146, -14.61, 17.02],
              'MG1048':[38.146, -10.562, 18.522],
              'MG1053':[38.146, -11.449, 22.008],
              'MG1151':[40.703, -15.986, 20.351],
              'MG1146':[40.703, -11.777, 22.669],
              'MG1148':[40.703, -13.409, 25.373],
              'MG1366':[43.316, -19.213, 25.02],
              'MG1350':[43.316, -13.546, 27.435],
              'MG1348':[43.316, -15.355, 31.017],
              'BG1548':[45.88, -13.519, 27.439],
              'BG1542':[45.88, -15.449, 30.991],
              'BG1544':[45.88, -16.916, 34.674],
              'BG1748':[48.448, -15.484, 30.905],
              'BG1743':[48.448, -16.958, 34.49],
              'BG1952':[50.978, -16.624, 34.363],
              'BG1955':[50.978, -13.52, 35.343],
              'BE0920':[35.538, 1.56, 11.939],
              'BE1225':[40.751, -0.638, 15.773],
              'BE1230':[40.751, 1.841, 12.319],
              'BE1528':[45.88, -0.581, 19.228],
              'BE1532':[45.88, 1.568, 15.512],
              'BE1824':[48.958, -0.873, 19.084],
              'BE1832':[48.958, 1.527, 15.373],
              'BE1928':[50.978, -1.116, 18.883],
              'BE1932':[50.978, 2.341, 19.65],
              'BE1935':[50.978, 4.142, 13.964],
              'RE1025':[38.656, 10.005, 14.867],
              'RE1328':[42.808, 15.233, 18.384],
              'RE1630':[46.922, 11.844, 12.337],
              'RE1632':[46.922, 14.823, 17.992],
              'YE1932':[50.978, 10.007, 26.137],
              'YE1937':[50.978, 13.167, 22.369],
              'YE2337':[55.031, 10.635, 30.011],
              'YE2344':[55.031, 13.063, 22.276],
              'YE2735':[59.067, 10.698, 29.749],
              'YE2740':[59.067, 12.998, 22.111],
              'YE3242':[63.011, 5.958, 31.797],
              'YE3245':[63.011, 10.535, 29.511],
              'YE3250':[63.011, 12.058, 37.976],
              'YE3755':[66.976, 6.043, 38.723],
              'YE3760':[66.976, 12.971, 42.748],
              'SE1932':[50.978, 4.193, 17.536], 
              'SE2335':[55.031, 5.348, 24.6],
              'SE2332':[55.031, 4.306, 17.658],
              'SE2740':[59.067, 5.314, 24.639],
              'SE2735':[59.067, 5.395, 16.822],
              'SE3445':[64.963, 1.838, 17.919],
              'SE3440':[64.963, 4.56, 13.035],
              'SE3945':[68.896, 2.123, 12.069],
              'SE3948':[68.896, 4.791, 17.37],
              'NB0407':[25.26, 1.158, -0.449],
              'NB0609':[30.40, 1.293, -0.479],
              'NB0911':[35.54, 1.472, -0.515],
              'NG1214':[40.751, 1.625, -0.58],
              'NG1517':[45.88, 1.751, -0.646],
              'NG1922':[50.978, 1.927, -0.682],
              'NG2427':[56.028, 2.044, -0.742],
              'NG2933':[61.045, 2.208, -0.784],
              'NG3538':[66.059, 2.343, -0.84],
              'NG4247':[70.871, 2.49, -0.888],
              'NG4954':[75.638, 2.611, -0.941],
              'NG5862':[80.516, 2.765, -0.988],
              'NW6770':[85.403, 2.888, -1.046], 
              'NW7780':[90.229, 3.043, -1.09],
              'NW8889':[95.077, 3.178, -1.137],
              'WR1216':[40.80, -7.59, 8.16],
              'WR2937':[61.05, -12.60, -5.57],
              'WR4250':[70.87, 4.97, 13.32],
              'WO0911':[35.52, -3.38, -12.29],
              'WO5363':[78.05, -10.48, -7.61],
              'CB1965':[50.978, -4.713, -16.964],
              'CB2980':[61.045, -12.131, -37.265],
              'CB4382':[71.214, -25.198, -25.499],
              'CY4970':[75.638, 18.525, 31.211],
              'CY6780':[85.352, 9.846, 18.865],
              'CY7785':[90.229, -5.577, 25.272],
              'CR1958':[50.978, 45.318, 32.692],
              'CR2260':[53.492, 44.555, 18.778],
              'CR2964':[61.045, 22.215, 7.098]}

data = []
label = []

for key, value in std_colors.items():
    data.append(value)
    label.append(key)
    pass

data = np.array(data)

# knn
def colour_classify(incolour, traindata, traincolour, k):
    '''训练的颜色个数'''
    #shape为numpy模块中的方法 shape[0]为矩阵第二维的长度
    trainsize = traindata.shape[0]
    #计算各个维度的差值并储存在向量diffmat中
    diffmat = np.tile(incolour, (trainsize,1)) - traindata
    #计算误差的平方
    squarediffmat = diffmat**2
    #计算向量间的欧式距离
    errordistance = squarediffmat.sum(axis=1)**0.5
    #排序
    sorteddistance = errordistance.argsort()
    classcount = {}
    for i in range(k):
        #选取前k个最符合要求的颜色
        selectedcolour = traincolour[sorteddistance[i]]
        classcount[selectedcolour] = classcount.get(selectedcolour,0)+1
        pass
    sortedclasscount = sorted(classcount.items(),
                              key=operator.itemgetter(1),reverse=True)
    return sortedclasscount[0][0]

def select_std_color(incolour):
    return colour_classify(incolour, data, label, 1)


class UI(object):
    
    def __init__(self):
        self.root = tk.Tk()
        MainFace(self.root)
        pass
    
    def run(self):
        self.root.mainloop()
        pass
    
    pass


class MainFace(object):
    
    def __init__(self, root):
        self.root = root
        self.root.title('Colour Selection Algorithm')
        self.root.geometry('860x780')
        self.face = tk.Frame(self.root)
        self.face.config(height=780, width=860)
        self.face.place(x=0, y=0)
        button1 = tk.Button(self.face, text='背景色选取', bg='#D2F99A', \
                            width=20, height=8, \
                            command=self.change1)
        button1.place(x=160, y=320)
        button2 = tk.Button(self.face, text='多背景共有色选取', bg='#F9CBF0', \
                            width=20, height=8, \
                            command=self.change2)
        button2.place(x=560, y=320)
        label = tk.Label(self.face, text='多背景共有色选取软件', \
                      font=('Arial', 20), width=32, height=2, justify=tk.LEFT)
        label.place(x=180, y=100)
        pass
    
    def change1(self):
        self.face.destroy()
        SingleFace(self.root)
        pass
    
    def change2(self):
        self.face.destroy()
        DoubleFace(self.root)
        pass
    
    pass


class SingleFace(object):
    
    def __init__(self, root):
        self.root = root
        self.face = tk.Frame(self.root)
        self.face.config(height=780, width=860)
        self.face.place(x=0, y=0)    
        self.image_path = tk.StringVar()
        l1 = tk.Label(self.face, text='背景图片：', \
                      font=('Arial', 12), width=16, height=2, justify=tk.LEFT)
        l1.place(x=2+5)
        self.e1 = tk.Entry(self.face, textvariable=self.image_path, width=72)
        self.e1.place(x=190+5, y=15)
        b1 = tk.Button(self.face, text='选择文件', width=8, height=1, command=self.select_path)
        b1.place(x=720+5, y=8) 
        self.lp1 = tk.Label(self.face, width=128, height=64)
        self.lp1.place(x=7, y=208)
        self.lp1c = tk.Label(self.face, width=640, height=64)
        self.lp1c.place(x=160, y=208)
        self.state = tk.StringVar() 
        self.state.set('正常')
        l2 = tk.Label(self.face, text='状态：', \
                      font=('Arial', 12), width=8, height=2, justify=tk.LEFT)
        l2.place(x=220, y=153)
        self.lstate = tk.Label(self.face, textvariable=self.state, \
                               font=('Arial', 12), width=50, height=1, justify=tk.LEFT, bg='#7FFF7F')
        self.lstate.place(x=285, y=162)
        l3 = tk.Label(self.face, text='聚类阈值：', \
                      font=('Arial', 12), width=16, height=2, justify=tk.LEFT)
        l3.place(x=7, y=100)
        self.e2 = tk.Entry(self.face, width=5)
        self.e2.place(x=190+5, y=115)
        self.scrollbar = tk.Scrollbar(self.face)
        self.scrollbar.place(x=808, y=368, height=368)
        self.listbox = tk.Listbox(self.face, yscrollcommand=self.scrollbar.set, \
                                  width=108, height=20)
        self.listbox.place(x=48, y=368)
        themes_pic_button = tk.Button(self.face, text='改变颜色条宽度', \
                                      width=16, height=2, \
                                      command=self.change_color_bar)
        themes_pic_button.place(x=370, y=300)
        back_button = tk.Button(self.face, text='返回主菜单', \
                                  width=10, height=1, \
                                  command=self.back)
        back_button.place(x=495, y=740)
        save_button = tk.Button(self.face, text='导出数据', \
                                  width=10, height=1, \
                                  command=self.save_excel)
        save_button.place(x=295, y=740)
        b2 = tk.Button(self.face, text='开始选取', width=25, height=1, \
                       bg='#7F7FFF', command=self.main) 
        b2.place(x=7, y=158) 
        pass
    
    def select_path(self):
        path = askopenfilename()
        self.image_path.set(path)
        pass
    
    def change_color_bar(self):
        try:
            self.make_themes_image_width = self.make_themes_image_width // 2
            if self.make_themes_image_width < 4:
                self.make_themes_image_width = 32
                pass
            pass
        except:
            pass
        else:
            pic_array = self.make_themes_image(self.themes_pic_array, \
                                               self.make_themes_image_width)
            pic = Image.fromarray(pic_array.astype('uint8')).convert('RGB')
            img = ImageTk.PhotoImage(pic)
            self.lp1c.config(image=img)
            self.lp1c.image = img
            self.face.update_idletasks()
            pass
        pass
    
    def make_themes_image(self, theme, width=32):
        self.make_themes_image_width = width
        output = [[]]
        for pix in theme:
            for i in range(width):
                output[0].append(pix)
                pass
            pass
        for i in range(64):
            output.append(output[0])
            pass
        output = np.array(output)
        return output
    
    def main(self):
        # check input
        path = self.e1.get()
        try:
            image_RGB = plt.imread(path)
            cluster_tolerance = float(self.e2.get())
            pass
        except:
            self.state.set('ERROR')
            self.lstate.config(bg='#FF7F7F')
            self.face.update_idletasks()
            messagebox.showinfo(title='ERROR', message='输入错误!')
            return None
        self.lstate.config(bg='#7FFF7F')
        self.state.set('正常')
        self.face.update_idletasks()
        # show image
        self.state.set('显示图片中。。。')
        self.face.update_idletasks()
        img_open = Image.open(path)
        img = img_open.resize((128, 64))
        img = ImageTk.PhotoImage(img)
        self.lp1.config(image=img)
        self.lp1.image = img
        self.face.update_idletasks()
        # resize to array
        image_RGB = Image.open(path)
        w_resize = 96
        h_resize = int(w_resize*image_RGB.size[1]/image_RGB.size[0])
        image_RGB = image_RGB.resize((w_resize, h_resize))
        image_RGB = np.array(image_RGB)
        # to lab
        self.state.set('转换RGB为LAB中。。。')
        self.face.update_idletasks()
        image_LAB = cv2.cvtColor(image_RGB, cv2.COLOR_RGB2LAB)
        # cluster
        self.state.set('图片聚类中。。。')
        self.face.update_idletasks()
        dbscan = DBSCAN(eps=cluster_tolerance, min_samples=1)
        h_1, w_1, c_1 = image_LAB.shape
        image_data = image_LAB.reshape((h_1*w_1, c_1))
        image_lab_data = []
        # uint8 to true lab
        for data in image_data:
            image_lab_data.append([data[0]*100/255, data[1]-128, data[2]-128])
            pass
        image_lab_data = np.array(image_lab_data)
        dbscan.fit(image_lab_data)
        labels = dbscan.labels_
        n_clusters = len(set(labels)) - (1 if -1 in labels else 0)
        # find the cluster center
        themes = []
        clusters_area = []
        for i in range(n_clusters):
            one_cluster = image_lab_data[labels == i]
            if len(one_cluster)!=1:
                km = KMeans(n_clusters=1, max_iter=300)
                km.fit(one_cluster)
                themes.append(np.squeeze(km.cluster_centers_))
                pass
            else:
                themes.append(one_cluster[0])
                pass
            clusters_area.append(len(one_cluster)/len(image_lab_data))
            pass
        themes = np.array(themes)
        # show themes image
        uint8_themes = []
        for theme in themes:
            uint8_themes.append([theme[0]*255/100, theme[1]+128, theme[2]+128])
            pass
        uint8_themes = np.array(uint8_themes)
        pic_array = cv2.cvtColor(np.uint8(uint8_themes.reshape(1, len(uint8_themes), 3)), \
                                 cv2.COLOR_LAB2RGB)[0]
        self.themes_pic_array = pic_array
        pic_array = self.make_themes_image(pic_array)
        pic = Image.fromarray(pic_array.astype('uint8')).convert('RGB')
        img = ImageTk.PhotoImage(pic)
        self.lp1c.config(image=img)
        self.lp1c.image = img
        self.face.update_idletasks()
        self.state.set('聚类完成')
        self.face.update_idletasks()
        themes_areas = list(zip(themes, clusters_area))
        # find std
        themes_areas_stds = []
        for theme, area in themes_areas:
            std = select_std_color(theme)
            themes_areas_stds.append((theme, area, std))
            pass
        # compute std area
        sorted_themes_areas_stds = sorted(themes_areas_stds, key=lambda x:(x[2]))
        last_std = None
        std_area = 0
        std_area_dict = {}
        for theme, area, std in sorted_themes_areas_stds:
            if std == last_std:
                std_area += area
                pass
            else:
                std_area = area
                pass
            std_area_dict[std] = std_area
            last_std = std
            pass  
        # get chunk
        std_chunk_dict = {}
        for key, value in std_area_dict.items():
            for theme, area, std in themes_areas_stds:
                if key == std:
                    info = (theme, area, std, value)
                    try:
                        std_chunk_dict[key].append(info)
                        pass
                    except:
                        std_chunk_dict[key] = [info]
                        pass
                    pass
                pass
            pass
        # sort
        results = []
        sorted_stds_areas = sorted(list(std_area_dict.items()), key=lambda x:(x[1]), reverse=True)
        for std, area in sorted_stds_areas:
            infos = std_chunk_dict[std]
            sorted_infos = sorted(infos, key=lambda x:(x[1]), reverse=True)
            for info in sorted_infos:
                results.append(info)
                pass
            pass
        # clear up
        self.listbox.delete(0, tk.END)
        self.face.update_idletasks()
        self.state.set('处理列表中。。。请稍后')
        self.face.update_idletasks()
        # for excel
        self.results = []
        # write to list
        info = ' '*(8-len('序号')) + str('序号')
        info += ' '*(45-len('LAB值')) + str('LAB值')
        info += ' '*(30-len('面积占比')) + str('面积占比')
        info += ' '*(25-len('军标色')) + str('军标色')
        info += ' '*(30-len('面积占比')) + str('面积占比')
        self.listbox.insert(tk.END, info)
        self.face.update_idletasks()
        # index
        index = 0
        last_std = None
        for theme, area, std, std_area in results:
            index = index + 1
            L, A, B = theme
            L = round(L, 3)
            A = round(A, 3)
            B = round(B, 3)
            result = (index, [L, A, B], round(100*area, 3), std, round(100*std_area, 3))
            self.results.append(result)
            info = ' '*(8-len(str(result[0]))) + str(result[0])
            info += ' '*(45-len(str(result[1]))) + str(result[1])
            info += ' '*(30-len(str(result[2]))) + str(result[2]) + '%'
            info += ' '*(25-len(str(result[3]))) + str(result[3])
            if last_std != std:
                info += ' '*(30-len(str(result[4]))) + str(result[4]) + '%'
                pass
            last_std = std
            self.listbox.insert(tk.END, info)
            self.face.update_idletasks()
            pass
        self.scrollbar.config(command=self.listbox.yview)
        self.face.update_idletasks()
        self.state.set('选取完成')
        self.face.update_idletasks()
        pass
        
    def save_excel(self):
        options = {}
        options['filetypes'] = [('Excel 文件', '.xlsx')]
        options['initialfile'] = 'untitled1.xlsx'
        save_path = asksaveasfilename(**options)
        try:
            wb = Workbook()
            sheet = wb.active
            sheet.column_dimensions['B'].width = 35.0
            sheet.column_dimensions['C'].width = 15.0
            sheet.column_dimensions['E'].width = 15.0
            sheet['A1'] = '序号'
            sheet['B1'] = 'LAB值'
            sheet['C1'] = '面积占比'
            sheet['D1'] = '军标色号'
            sheet['E1'] = '面积占比'
            # write infos
            last_std = None
            for result in self.results:
                index, lab, area, std, std_area = result
                sheet['A'+str(index+1)] = index
                sheet['B'+str(index+1)] = str(lab)
                sheet['C'+str(index+1)] = str(area) + '%'
                sheet['D'+str(index+1)] = str(std)
                if last_std != std:
                    sheet['E'+str(index+1)] = str(std_area) + '%'
                    pass
                last_std = std
                pass
            # cell style
            align = Alignment(horizontal='center', vertical='center')
            for i in range(1, len(self.results)+2):
                sheet['A'+str(i)].alignment = align
                sheet['B'+str(i)].alignment = align
                sheet['C'+str(i)].alignment = align
                sheet['D'+str(i)].alignment = align
                sheet['E'+str(i)].alignment = align
                pass
            pass
        except:
            self.state.set('ERROR')
            self.lstate.config(bg='#FF7F7F')
            self.face.update_idletasks()
            messagebox.showinfo(title='ERROR', message='请先进行图片选取!')
            pass
        else:
            try:
                wb.save(save_path)
                pass
            except:
                self.state.set('保存失败，文件可能被占用')
                self.lstate.config(bg='#FF7F7F')
                self.face.update_idletasks()
                pass
            else:
                # save
                self.lstate.config(bg='#7FFF7F')
                self.state.set('保存成功')             
                self.face.update_idletasks()
                pass
            pass
        pass
    
    def back(self):
        self.face.destroy()
        MainFace(self.root)
        pass
    
    pass


class DoubleFace(object):
    
    def __init__(self, root):
        self.root = root
        self.face = tk.Frame(self.root)
        self.face.config(height=780, width=860)
        self.face.place(x=0, y=0)
        l1 = tk.Label(self.face, text='背景1：', \
                      font=('Arial', 12), width=16, height=2, justify=tk.LEFT)
        l1.place(x=2+5)
        l2 = tk.Label(self.face, text='背景2：', \
                      font=('Arial', 12), width=16, height=2, justify=tk.LEFT)
        l2.place(x=2+5, y=50)
        self.path_1 = tk.StringVar()
        self.e1 = tk.Entry(self.face, textvariable=self.path_1, width=72)
        self.e1.place(x=190+5, y=15)
        self.path_2 = tk.StringVar()
        self.e2 = tk.Entry(self.face, textvariable=self.path_2, width=72)
        self.e2.place(x=190+5, y=65)
        b1 = tk.Button(self.face, text='选择文件', width=8, height=1, \
                       command=self.select_path_1)
        b1.place(x=720+5, y=8) 
        b2 = tk.Button(self.face, text='选择文件', width=8, height=1, \
                       command=self.select_path_2) 
        b2.place(x=720+5, y=58) 
        self.lp1 = tk.Label(self.face, width=128, height=64)
        self.lp1.place(x=7, y=208)
        self.lp2 = tk.Label(self.face, width=128, height=64)
        self.lp2.place(x=7, y=288)
        self.lp1c = tk.Label(self.face, width=640, height=64)
        self.lp1c.place(x=160, y=208)
        self.lp2c = tk.Label(self.face, width=640, height=64)
        self.lp2c.place(x=160, y=288)
        l3 = tk.Label(self.face, text='聚类阈值：', \
                      font=('Arial', 12), width=16, height=2, justify=tk.LEFT)
        l3.place(x=7, y=100)
        self.e3 = tk.Entry(self.face, width=5)
        self.e3.place(x=190+5, y=115)
        l4 = tk.Label(self.face, text='色差阈值：', \
                      font=('Arial', 12), width=16, height=2, justify=tk.LEFT)
        l4.place(x=360, y=100)
        self.e4 = tk.Entry(self.face, width=5)
        self.e4.place(x=360+183+5, y=115)
        self.state = tk.StringVar() 
        self.state.set('正常')
        l5 = tk.Label(self.face, text='状态：', \
                      font=('Arial', 12), width=8, height=2, justify=tk.LEFT)
        l5.place(x=220, y=153)
        self.lstate = tk.Label(self.face, textvariable=self.state, \
                               font=('Arial', 12), width=50, height=1, justify=tk.LEFT, bg='#7FFF7F')
        self.lstate.place(x=285, y=162)
        self.scrollbar = tk.Scrollbar(self.face)
        self.scrollbar.place(x=808, y=368, height=368)
        self.listbox = tk.Listbox(self.face, yscrollcommand=self.scrollbar.set, width=108, height=20)
        self.listbox.place(x=48, y=368)
        b3 = tk.Button(self.face, text='开始选取', width=25, height=1, bg='#7F7FFF', \
                       command=self.main) 
        b3.place(x=7, y=158) 
        back_button = tk.Button(self.face, text='返回主菜单', \
                                  width=10, height=1, \
                                  command=self.back)
        back_button.place(x=495, y=740)
        save_button = tk.Button(self.face, text='导出数据', \
                                  width=10, height=1, \
                                  command=self.save_excel)
        save_button.place(x=295, y=740)
        pass
    
    def select_path_1(self):
        path = askopenfilename()
        self.path_1.set(path)
        pass

    def select_path_2(self):
        path = askopenfilename()
        self.path_2.set(path)
        pass
    
    def make_themes_image(self, theme, width=32):
        self.make_themes_image_width = width
        output = [[]]
        for pix in theme:
            for i in range(width):
                output[0].append(pix)
                pass
            pass
        for i in range(64):
            output.append(output[0])
            pass
        output = np.array(output)
        return output
    
    def calc_chromatism(self, lab1, lab2):
        deltaL = lab1[0] - lab2[0]
        deltaA = lab1[1] - lab2[1]
        deltaB = lab1[2] - lab2[2]
        deltaE = (deltaL**2 + deltaA**2 + deltaB**2)**0.5
        return deltaE
    
    def main(self):    
        image_1_path = self.e1.get()
        image_2_path = self.e2.get()
        try:
            image_1_RGB = plt.imread(image_1_path)
            image_2_RGB = plt.imread(image_2_path)
            cluster_tolerance = float(self.e3.get())
            color_tolerance = float(self.e4.get())
            pass
        except:
            self.state.set('ERROR')
            self.lstate.config(bg='#FF7F7F')
            self.face.update_idletasks()
            messagebox.showinfo(title='ERROR', message='输入错误!')
            return None
        self.lstate.config(bg='#7FFF7F')
        self.face.update_idletasks()
        # show images
        self.state.set('显示图片中。。。')
        self.face.update_idletasks()
        img_open = Image.open(self.e1.get())
        img = img_open.resize((128, 64))
        img = ImageTk.PhotoImage(img)
        self.lp1.config(image=img)
        self.lp1.image = img
        self.face.update_idletasks()
        img_open = Image.open(self.e2.get())
        img = img_open.resize((128, 64))
        img = ImageTk.PhotoImage(img)
        self.lp2.config(image=img)
        self.lp2.image = img
        self.face.update_idletasks()
        # resize to speed up
        image_1_RGB = Image.open(image_1_path)
        w_resize = 96
        h_resize = int(w_resize*image_1_RGB.size[1]/image_1_RGB.size[0])
        image_1_RGB = image_1_RGB.resize((w_resize, h_resize))
        image_1_RGB = np.array(image_1_RGB)
        # resize to speed up
        image_2_RGB = Image.open(image_2_path)
        w_resize = 96
        h_resize = int(w_resize*image_2_RGB.size[1]/image_2_RGB.size[0])
        image_2_RGB = image_2_RGB.resize((w_resize, h_resize))
        image_2_RGB = np.array(image_2_RGB)
        # to lab
        self.state.set('转换RGB为LAB中。。。')
        self.face.update_idletasks()
        image_1_LAB = cv2.cvtColor(image_1_RGB, cv2.COLOR_RGB2LAB)
        image_2_LAB = cv2.cvtColor(image_2_RGB, cv2.COLOR_RGB2LAB)  
        # image 1
        self.state.set('第一张图片聚类中。。。')
        self.face.update_idletasks()
        dbscan1 = DBSCAN(eps=cluster_tolerance, min_samples=1)
        h_1, w_1, c_1 = image_1_LAB.shape
        image_1_data = image_1_LAB.reshape((h_1*w_1, c_1))
        image_1_lab_data = []
        for data in image_1_data:
            image_1_lab_data.append([data[0]*100/255, data[1]-128, data[2]-128])
            pass
        image_1_lab_data = np.array(image_1_lab_data)
        dbscan1.fit(image_1_lab_data)
        labels = dbscan1.labels_
        n_clusters_1 = len(set(labels)) - (1 if -1 in labels else 0)
        # find the cluster center
        themes_1 = []
        clusters_area_1 = []
        for i in range(n_clusters_1):
            one_cluster = image_1_lab_data[labels == i]
            if len(one_cluster)!=1:
                km = KMeans(n_clusters=1, max_iter=300)
                km.fit(one_cluster)
                themes_1.append(np.squeeze(km.cluster_centers_))
                pass
            else:
                themes_1.append(one_cluster[0])
                pass
            clusters_area_1.append(len(one_cluster)/len(image_1_lab_data))
            pass
        themes_1 = np.array(themes_1)
        # show image
        uint8_themes_1 = []
        for theme in themes_1:
            uint8_themes_1.append([theme[0]*255/100, theme[1]+128, theme[2]+128])
            pass
        uint8_themes_1 = np.array(uint8_themes_1)   
        pic_array = cv2.cvtColor(np.uint8(uint8_themes_1.reshape(1, len(uint8_themes_1), 3)), \
                                 cv2.COLOR_LAB2RGB)
        pic_array = self.make_themes_image(pic_array[0])
        pic = Image.fromarray(pic_array.astype('uint8')).convert('RGB')
        img = ImageTk.PhotoImage(pic)
        self.lp1c.config(image=img)
        self.lp1c.image = img
        self.face.update_idletasks()
        # image 2
        self.state.set('第二张图片聚类中。。。')
        self.face.update_idletasks()
        dbscan2 = DBSCAN(eps=cluster_tolerance, min_samples=1)
        h_2, w_2, c_2 = image_2_LAB.shape
        image_2_data = image_2_LAB.reshape((h_2*w_2, c_2))
        image_2_lab_data = []
        for data in image_2_data:
            image_2_lab_data.append([data[0]*100/255, data[1]-128, data[2]-128])
            pass
        image_2_lab_data = np.array(image_2_lab_data)
        dbscan2.fit(image_2_lab_data)
        labels = dbscan2.labels_
        n_clusters_2 = len(set(labels)) - (1 if -1 in labels else 0)
        # find the cluster center
        themes_2 = []
        clusters_area_2 = []
        for i in range(n_clusters_2):
            one_cluster = image_2_lab_data[labels == i]
            if len(one_cluster)!=1:
                km = KMeans(n_clusters=1, max_iter=300)
                km.fit(one_cluster)
                themes_2.append(np.squeeze(km.cluster_centers_))
                pass
            else:
                themes_2.append(one_cluster[0])
                pass
            clusters_area_2.append(len(one_cluster)/len(image_2_lab_data))
            pass
        themes_2 = np.array(themes_2)
        # show image
        uint8_themes_2 = []
        for theme in themes_2:
            uint8_themes_2.append([theme[0]*255/100, theme[1]+128, theme[2]+128])
            pass
        uint8_themes_2 = np.array(uint8_themes_2)
        pic_array = cv2.cvtColor(np.uint8(uint8_themes_2.reshape(1, len(uint8_themes_2), 3)), \
                                 cv2.COLOR_LAB2RGB)
        pic_array = self.make_themes_image(pic_array[0])
        pic = Image.fromarray(pic_array.astype('uint8')).convert('RGB')
        img = ImageTk.PhotoImage(pic)
        self.lp2c.config(image=img)
        self.lp2c.image = img
        self.face.update_idletasks()
        self.state.set('聚类完成')
        self.face.update_idletasks()
        # select common color
        Image_1_Area = clusters_area_1[:]
        Image_2_Area = clusters_area_2[:]
        self.state.set('共同色选取中。。。')
        self.face.update_idletasks()
        common_color_infos = []
        for i in range(n_clusters_1):
            L1 = themes_1[i][0]
            A1 = themes_1[i][1]
            B1 = themes_1[i][2]
            LAB1 = [L1, A1, B1]
            for j in range(n_clusters_2):
                L2 = themes_2[j][0]
                A2 = themes_2[j][1]
                B2 = themes_2[j][2]
                LAB2 = [L2, A2, B2]
                deltaE = self.calc_chromatism(LAB1, LAB2)
                if deltaE <= color_tolerance:
                    S1 = Image_1_Area[i] / (Image_1_Area[i] + Image_2_Area[j])
                    S2 = Image_2_Area[j] / (Image_1_Area[i] + Image_2_Area[j])
                    L3 = L1 * S1 + L2 * S2
                    A3 = A1 * S1 + A2 * S2
                    B3 = B1 * S1 + B2 * S2
                    L1 = round(L1, 3)
                    A1 = round(A1, 3)
                    B1 = round(B1, 3)
                    L2 = round(L2, 3)
                    A2 = round(A2, 3)
                    B2 = round(B2, 3)
                    L3 = round(L3, 3)
                    A3 = round(A3, 3)
                    B3 = round(B3, 3)
                    LAB1 = [L1, A1, B1]
                    LAB2 = [L2, A2, B2]
                    LAB3 = [L3, A3, B3]
                    selected_std_color = select_std_color(LAB3)
                    selected_std_color_lab = std_colors[selected_std_color]
                    # knn
                    label = colour_classify(selected_std_color_lab, \
                                            image_1_lab_data, np.squeeze(dbscan1.labels_), k=10)
                    # area
                    std_color_area_1 = clusters_area_1[label]
                    # knn
                    label = colour_classify(selected_std_color_lab, \
                                            image_2_lab_data, np.squeeze(dbscan2.labels_), k=10)
                    # area
                    std_color_area_2 = clusters_area_2[label]
                    # info
                    info = (LAB3, LAB1, Image_1_Area[i], LAB2, Image_2_Area[j], \
                            selected_std_color, std_color_area_1, std_color_area_2)
                    common_color_infos.append(info)
                    pass
                pass
            pass
        self.state.set('共同色选取完成')
        self.face.update_idletasks()
        selected_std_colors = []
        # keys: num appears values: selected std colors
        dict_selected_std_colors = {}
        # num appears
        std_colors_nums = []
        for i in range(len(common_color_infos)):
            # std color index: -3
            selected_std_colors.append(common_color_infos[i][-3])
            selected_std_colors_set = set(selected_std_colors)
            pass
        for selected_std_color in selected_std_colors_set:
            num = selected_std_colors.count(selected_std_color)
            if str(num) not in dict_selected_std_colors.keys():  
                std_colors_nums.append(num)
                dict_selected_std_colors[str(num)] = [selected_std_color]
                pass
            else:
                dict_selected_std_colors[str(num)].append(selected_std_color)
                pass
            pass
        std_colors_nums.sort(reverse=True)
        # list box
        index = 0
        self.listbox.delete(0, tk.END)
        self.face.update_idletasks()
        info = ' '*2 + str('') + ' '*(7-len(str('')))
        info += ' '*8 + str('') + ' '*(17-len(str('')))
        info += ' '*12
        info += ' '*15 + str('背景1') + ' '*(20-len(str('背景1')))
        info += ' '*15 + str('背景2') + ' '*(20-len(str('背景2')))
        info += ' '*16 + str('军标色') + ' '*(14-len(str('军标色'))) 
        self.listbox.insert(tk.END, info)
        self.face.update_idletasks()   
        info = ' '*1 + str('序号') + ' '*(8-len(str('序号')))
        info += ' '*8 + str('共同色') + ' '*(17-len(str('共同色')))
        info += ' '*10 + str('LAB1') + ' '*(15-len(str('LAB1')))
        info += ' '*5 + str('Area1') + ' '*(5-len(str('Area1')))
        info += ' '*10 + str('LAB2') + ' '*(15-len(str('LAB2')))
        info += ' '*5 + str('Area2') + ' '*(5-len(str('Area2')))
        info += ' '*6 + str('色号') + ' '*(4-len(str('色号'))) 
        info += ' '*5 + str('背景1占比') + ' '*(5-len(str('背景1占比'))-1)
        info += ' '*5 + str('背景2占比') + ' '*(5-len(str('背景2占比'))-1)
        self.listbox.insert(tk.END, info)
        self.face.update_idletasks()
        self.state.set('处理列表中。。。请稍后')
        self.face.update_idletasks()
        self.results = []
        for num in std_colors_nums:
            for color in dict_selected_std_colors[str(num)]:
                count = 0
                for color_info in common_color_infos:
                    # std color index: -3
                    if color_info[-3] == color:
                        index += 1
                        count += 1
                        c3, c1, a1, c2, a2, \
                        sc, sca1, sca2 = color_info
                        a1 = round(100*a1, 3)
                        a2 = round(100*a2, 3)
                        sca1 = round(100*sca1, 3)
                        sca2 = round(100*sca2, 3)
                        if count<=1:
                            self.results.append((index, c3, c1, a1, c2, a2, sc, \
                                                sca1, sca2))
                            pass
                        else:
                            self.results.append((index, c3, c1, a1, c2, a2, sc, \
                                                None, None))
                            pass
                        info = ' '*2 + str(index) + ' '*(7-len(str(index)))
                        info += str(c3) + ' '*(25-len(str(c3)))
                        info += str(c1) + ' '*(25-len(str(c1)))
                        info += str(a1) + '%' + ' '*(10-len(str(a1))-1)
                        info += str(c2) + ' '*(25-len(str(c2)))
                        info += str(a2) + '%' + ' '*(10-len(str(a2))-1)
                        info += str(sc) + ' '*(10-len(str(sc))) 
                        if count<=1:
                            info += str(sca1) + '%' + ' '*(10-len(str(sca1))-1)
                            info += str(sca2) + '%' + ' '*(10-len(str(sca2))-1)
                            pass
                        self.listbox.insert(tk.END, info)
                        self.face.update_idletasks()
                        pass
                    pass
                pass
            pass
        self.scrollbar.config(command=self.listbox.yview)
        self.face.update_idletasks()
        self.state.set('选取完成')
        self.face.update_idletasks()
        pass
    
    def save_excel(self):
        options = {}
        options['filetypes'] = [('Excel 文件', '.xlsx')]
        options['initialfile'] = 'untitled1.xlsx'
        save_path = asksaveasfilename(**options)
        try:
            wb = Workbook()
            sheet = wb.active
            sheet.merge_cells('A1:A2')
            sheet.merge_cells('B1:B2')
            sheet.merge_cells('C1:D1')
            sheet.merge_cells('E1:F1')
            sheet.merge_cells('G1:I1')
            sheet.cell(1,1).value = '序号'
            sheet.cell(1,2).value = '共同色'
            sheet.cell(1,3).value = '背景1'
            sheet.cell(1,5).value = '背景2'
            sheet.cell(1,7).value = '军标色'
            sheet.column_dimensions['B'].width = 35.0
            sheet.column_dimensions['C'].width = 35.0
            sheet.column_dimensions['E'].width = 35.0
            sheet.column_dimensions['D'].width = 15.0
            sheet.column_dimensions['F'].width = 15.0
            sheet.column_dimensions['H'].width = 15.0
            sheet.column_dimensions['I'].width = 15.0
            sheet['C2'] = 'LAB值'
            sheet['D2'] = '面积占比'
            sheet['E2'] = 'LAB值'
            sheet['F2'] = '面积占比'
            sheet['G2'] = '色号'
            sheet['H2'] = '背景1占比'
            sheet['I2'] = '背景2占比'
            for result in self.results:
                index, c3, c1, a1, c2, a2, \
                sc, sca1, sca2 = result
                sheet['A'+str(index+2)] = index
                sheet['B'+str(index+2)] = str(c3)
                sheet['C'+str(index+2)] = str(c1)
                sheet['D'+str(index+2)] = str(a1) + '%'
                sheet['E'+str(index+2)] = str(c2)
                sheet['F'+str(index+2)] = str(a2) + '%'
                sheet['G'+str(index+2)] = str(sc)
                if sca1 != None:
                    sheet['H'+str(index+2)] = str(sca1) + '%'
                    sheet['I'+str(index+2)] = str(sca2) + '%'
                    pass
                else:
                    sheet.merge_cells('H'+str(index+1)+":"+'H'+str(index+2))
                    sheet.merge_cells('I'+str(index+1)+":"+'I'+str(index+2))
                    pass
                pass
            # cell style
            align = Alignment(horizontal='center', vertical='center')
            for i in range(1, len(self.results)+3):
                sheet['A'+str(i)].alignment = align
                sheet['B'+str(i)].alignment = align
                sheet['C'+str(i)].alignment = align
                sheet['D'+str(i)].alignment = align
                sheet['E'+str(i)].alignment = align
                sheet['F'+str(i)].alignment = align
                sheet['G'+str(i)].alignment = align
                sheet['H'+str(i)].alignment = align
                sheet['I'+str(i)].alignment = align
                pass
            pass
        except:
            self.state.set('ERROR')
            self.lstate.config(bg='#FF7F7F')
            self.face.update_idletasks()
            messagebox.showinfo(title='ERROR', message='请先进行图片选取!')
            pass
        else:
            try:
                wb.save(save_path)
                pass
            except:
                self.state.set('保存失败，文件可能被占用')
                self.lstate.config(bg='#FF7F7F')
                self.face.update_idletasks()
                pass
            else:
                self.lstate.config(bg='#7FFF7F')
                self.state.set('保存成功')             
                self.face.update_idletasks()
                pass
            pass
        pass
    
    def back(self):
        self.face.destroy()
        MainFace(self.root)
        pass
    
    pass


if __name__ == '__main__':
    window = UI()
    window.run()
    pass
