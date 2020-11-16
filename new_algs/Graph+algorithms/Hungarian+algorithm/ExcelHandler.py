# -*- coding: utf-8 -*-
#coding:utf8

from openpyxl import load_workbook
from tkinter.filedialog import *
from tkinter.messagebox import *
import sys
import os
import datetime


def SetUp (_DataHandler):
    global DataHandler
    DataHandler = _DataHandler;
    LoadSettings(StartingSettingsFileName);


exportTemplate = "Data\ExportTemplate.xlsx"
exportFileName = "Exports\Export-"
StartingSettingsFileName = "Data\Settings.xlsx"


if getattr (sys, 'frozen', False):
    # frozen
    myPath = os.path.dirname (sys.executable)
else:
    # unfrozen
    myPath = os.path.dirname (os.path.realpath (__file__))


def LoadFromExternalExcelFile ():
    inputPath = askopenfilename (initialdir = "myPath", title = "Please Select a Jotfrom Generated Excel File");
    LoadFromExternalExcelFileWithPath (inputPath);


#load from an excel file generated with a jotform
def LoadFromExternalExcelFileWithPath (filePath):
    print("-*-*-");
    print("Loading from External Excel File")

    try:
        wb = load_workbook (filename = filePath)
        ws = wb.active
    except Exception as e:
        print(e)
        showerror("I/O Error", "Can't read the file. " + str(e));
        return


    print("Adding Students")
    x = 2
    while ws.cell(x,4).value != None:
        myName = ws.cell(x,4).value; #name
        myName += " " + ws.cell(x,5).value #surname
        stu = DataHandler.Student(grade = int(ws.cell(x,6).value), name = myName);

        stu.choices = []
        for n in range(DataHandler.periodCount):
            stu.choices.append([]);

        for period in range(DataHandler.periodCount):
            for act in ws.cell(x, 8 + period).value.split("\n"):
                if not (act == ""):
                    act = act.split("-");
                    act[0] = ""
                    actName = ""
                    for part in act:
                        if not part == "":
                            actName += part.replace("-","") + "-"

                    actName = actName[:-1]
                    #print(act[0] + " - " + act[1] + " - " + act[2])
                    stu.choices[period].append(DataHandler.Student.Act(name=actName, isForced=False))

        stu.assigned = [None] * DataHandler.periodCount
        DataHandler.AddStudent(stu)

        x+=1
    print("-*-*-");

    print("Loaded from External Excel File Successfuly")
    print("-*-*-");


#load from our internal save
def LoadSettings (filePath):
    print("-*-*-");
    print("Loading Settings from Settings File")
    try:
        wb = load_workbook (filename = filePath)
        ws = wb.active
    except Exception as e:
        print(e)
        showerror("I/O Error", "Can't read the file. " + str(e));
        return

    DataHandler.periodCount = int (ws["B1"].value)

    print("-*-*-");
    print("GradeValues")
    for grad in ws["D1"].value.replace(" ", "").split(";"):
        print(grad)
        if not(grad == ""):
            grad = grad.split(",");
            print(str(len(DataHandler.gradeMultipliers)) + " - " + str(int(grad[0])) + " - " + str(float(grad[1])))
            DataHandler.gradeMultipliers[int(grad[0])] = float(grad[1])
    print("-*-*-");

    print("slotValues")
    n = 0
    for slot in ws["F1"].value.replace(" ", "").split(";"):
        print(slot)
        if not (slot == ""):
            slot = slot.split(",")
            DataHandler.slotMultipliers[n] = [0]*2
            DataHandler.slotMultipliers[n][0] = int(slot[0])
            DataHandler.slotMultipliers[n][1] = float(slot[1])

            n+= 1
            if(n > DataHandler.periodCount):
                break;
    print("-*-*-");
    print (DataHandler.periodCount)
    print (DataHandler.gradeMultipliers)
    print (DataHandler.slotMultipliers)

    print("Loaded from Settings File Successfuly")
    print("-*-*-");


#------------------------------------------------------------------------------------------------------ EXPORT
def ExporttoExcelFile ():
    print ("-*-*-");
    print ("Exporting Excel File")

    date = datetime.datetime.now ();
    dateStr = str (date.day) + "-" + str (date.month) + "-" + str (date.year) + " - " + str (date.hour) + "-" + str (date.minute);

    try:
        wb = load_workbook (filename = exportTemplate)
        ws_students = wb["Öğrenciler"]
        ws_activities = wb["Aktiviteler"]

        wb.save(exportFileName + dateStr + '.xlsx')
    except Exception as e:
        print(e)
        if (askretrycancel ("I/O Error", "There was an error while Exporting. " +
                                         "You may continue your work, but your results aren't saved. " +
                                         "To fix this issue please try closing any open files, checking your antivirus, or running with administrator rights fix this issue. "
                                        + str(e))):
            return ExporttoExcelFile ();
        else:
            return


    print ("-*-*-");
    print("Students")
    x = 2
    for stu in DataHandler.allStudents:
        ws_students.cell (x, 1).value = str(stu.grade);
        ws_students.cell (x, 2).value = stu.name;

        p = 0
        assignedVals = ""
        for act in stu.assigned:
            if act != None:
                ws_students.cell (x, 3 + p).value = act.name;
                p+=1
        x+=1;

    print ("-*-*-");
    print("Activities")
    y = 1
    for act in DataHandler.allActivities:
        ws_activities.cell(1,y).value = str(act.period+1)+ " - " + str(act.name);
        x = 2
        for assStu in act.assigned:
            ws_activities.cell(x,y).value = str(assStu.name);
            x+=1
        y+=1

    wb.save (exportFileName + dateStr + '.xlsx')
    print ("Exported Successfuly: " + exportFileName + dateStr + '.xlsx')
    print ("-*-*-");